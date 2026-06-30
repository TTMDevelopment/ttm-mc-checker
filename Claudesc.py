import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import threading
import time
import os
import re
from datetime import datetime
from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ────────────────────────────────────────────────────────────────
BASE_URL  = "https://motus.dot.gov/customer/{}/account"
COLUMNS   = [
    "DOT Number", "Legal Business Name", "Doing Business As (DBA)",
    "Principal Place of Business", "Mailing Address", "Business Telephone No.",
    "Duns & Bradstreet", "Form of Business", "State Incorporated",
    "Business Email", "Official Name", "Title",
    "Contact Telephone No.", "Contact Email"
]
HEADER_BG  = "1F4E78"
HEADER_FG  = "FFFFFF"
GRID_COLOR = "D9D9D9"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ─── Known field labels exactly as they appear on the page ───────────────────
# Order matters: we search for these in sequence down the page.
BUSINESS_FIELDS = [
    ("legal_name",  "Legal Business Name"),
    ("dba",         "Doing Business As (DBA) Name"),
    ("principal",   "Principal Place of Business"),
    ("mailing",     "Mailing Address"),
    ("phone",       "Business Telephone No."),
    ("duns",        "Duns & Bradstreet"),
    ("form",        "Form of Business"),
    ("state_inc",   "State Incorporated"),
    ("email",       "Business Email"),
]

# The four officials column headers that appear as a consecutive block
OFFICIALS_HEADERS = ["Official Name", "Title", "Telephone No", "Email"]

# ─── Page text utilities ──────────────────────────────────────────────────────

def get_lines(page):
    """Return cleaned, non-empty lines from the full page innerText."""
    raw = page.evaluate("() => document.body.innerText")
    return [ln.strip() for ln in raw.split("\n") if ln.strip()]


def extract_business_fields(lines, log):
    """
    Walk lines sequentially.  When we see a known label, the NEXT line
    that is NOT itself a known label is the value.  Blank/empty values
    on the page appear as the next label immediately — we handle that by
    storing "" (empty string) and moving on.
    """
    all_labels_lower = {lbl.lower() for _, lbl in BUSINESS_FIELDS}
    # Also add officials headers so we don't treat them as values
    officials_lower  = {h.lower() for h in OFFICIALS_HEADERS}
    skip_lower       = all_labels_lower | officials_lower

    label_lookup = {lbl.lower(): key for key, lbl in BUSINESS_FIELDS}

    result = {}
    i = 0
    while i < len(lines):
        ll = lines[i].lower()
        if ll in label_lookup:
            key = label_lookup[ll]
            # Look ahead for a value
            value = ""
            for j in range(i + 1, min(i + 4, len(lines))):
                candidate = lines[j]
                if candidate.lower() in skip_lower:
                    break          # Next label — this field is blank
                value = candidate
                break
            result[key] = value
        i += 1

    log(f"[✦] Business fields: {result}")
    return result


def extract_officials(lines, log):
    """
    The page emits officials data in two consecutive blocks:
        Official Name
        Title
        Telephone No
        Email
        BRANDON FLORES       ← values start here
        OWNER
        (phone value or blank)
        (email value or blank)

    Strategy: find the index where all four header strings appear
    consecutively, then read the values that follow the last header.
    """
    # Normalize for searching
    hdr_lower = [h.lower() for h in OFFICIALS_HEADERS]
    n = len(hdr_lower)

    block_start = None
    for i in range(len(lines) - n + 1):
        window = [lines[i + k].lower() for k in range(n)]
        if window == hdr_lower:
            block_start = i
            break

    if block_start is None:
        # Fallback: find any partial run of the headers
        log("[Officials] Exact block not found — trying partial match")
        for i, line in enumerate(lines):
            if line.lower() == hdr_lower[0]:
                block_start = i
                break

    if block_start is None:
        log("[Officials] ⚠ Header block not found in page text")
        return {"name": "N/A", "title": "N/A", "phone": "N/A", "email": "N/A"}

    # Values start immediately after the last header in the block
    val_start = block_start + n
    log(f"[Officials] Header block at line {block_start}, values at {val_start}")
    log(f"[Officials] Value lines: {lines[val_start:val_start+6]}")

    def safe_val(offset):
        idx = val_start + offset
        if idx < len(lines):
            v = lines[idx]
            # Reject footer/pagination noise
            if re.search(r"row(s)? (per page|selected)|^\d+[-–]\d+", v.lower()):
                return "N/A"
            return v if v else "N/A"
        return "N/A"

    result = {
        "name":  safe_val(0),
        "title": safe_val(1),
        "phone": safe_val(2),
        "email": safe_val(3),
    }
    log(f"[Officials] ✓ {result}")
    return result


def ensure_officials_open(page, log):
    """Click the Company Officials accordion only if it is currently closed."""
    try:
        panels = page.query_selector_all("mat-expansion-panel")
        for panel in panels:
            header = panel.query_selector("mat-expansion-panel-header")
            if not header:
                continue
            if "official" not in header.inner_text().lower():
                continue
            cls  = header.get_attribute("class") or ""
            aria = header.get_attribute("aria-expanded") or ""
            if "mat-expanded" in cls or aria == "true":
                log("[✓] Officials panel already open")
            else:
                log("[↕] Opening officials panel")
                header.click()
                time.sleep(1.5)
            return
        # No labelled panel found — try the first expansion header
        header = page.query_selector("mat-expansion-panel-header")
        if header:
            cls  = header.get_attribute("class") or ""
            aria = header.get_attribute("aria-expanded") or ""
            if "mat-expanded" not in cls and aria != "true":
                log("[↕] Opening first expansion panel")
                header.click()
                time.sleep(1.5)
    except Exception as e:
        log(f"[!] Panel check error: {e}")


# ─── Main scrape ──────────────────────────────────────────────────────────────

def scrape_dot(page, dot, log):
    url = BASE_URL.format(dot)
    log(f"[→] {url}")
    page.goto(url, wait_until="networkidle", timeout=60000)
    time.sleep(2.5)

    ensure_officials_open(page, log)
    time.sleep(1)

    lines = get_lines(page)
    log(f"[✦] Total lines captured: {len(lines)}")

    biz      = extract_business_fields(lines, log)
    officials = extract_officials(lines, log)

    def g(key):
        v = biz.get(key, "")
        return v if v else "N/A"

    return {
        "DOT Number":                  dot,
        "Legal Business Name":         g("legal_name"),
        "Doing Business As (DBA)":     g("dba"),
        "Principal Place of Business": g("principal"),
        "Mailing Address":             g("mailing"),
        "Business Telephone No.":      g("phone"),
        "Duns & Bradstreet":           g("duns"),
        "Form of Business":            g("form"),
        "State Incorporated":          g("state_inc"),
        "Business Email":              g("email"),
        "Official Name":               officials["name"],
        "Title":                       officials["title"],
        "Contact Telephone No.":       officials["phone"],
        "Contact Email":               officials["email"],
    }


# ─── Pipeline ─────────────────────────────────────────────────────────────────

def run_pipeline(dot_numbers, filepath, headless, log):
    records = []
    log(f"[⚡] Pipeline start — {len(dot_numbers)} DOT number(s)")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)
        ctx     = browser.new_context(user_agent=USER_AGENT)
        page    = ctx.new_page()

        for i, dot in enumerate(dot_numbers, 1):
            dot = dot.strip()
            if not dot:
                continue
            log(f"\n── [{i}/{len(dot_numbers)}] DOT {dot} ──")
            try:
                records.append(scrape_dot(page, dot, log))
                log(f"[✓] DOT {dot} complete")
            except Exception as e:
                log(f"[✗] DOT {dot} error: {e}")
                empty = {col: "N/A" for col in COLUMNS}
                empty["DOT Number"] = dot
                records.append(empty)

        browser.close()

    if records:
        write_excel(records, filepath, log)
        log(f"\n[✅] Done — {len(records)} record(s) exported.")
    else:
        log("[⚠] No records to export.")


# ─── Excel writer ─────────────────────────────────────────────────────────────

def write_excel(records, filepath, log):
    df = pd.DataFrame(records, columns=COLUMNS)
    df.to_excel(filepath, index=False, sheet_name="DOT Intelligence")

    wb = load_workbook(filepath)
    ws = wb.active

    thin   = Side(style="thin", color=GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    hdr_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)

    for ci in range(1, len(COLUMNS) + 1):
        cell           = ws.cell(row=1, column=ci)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    CENTER_COLS = {
        "DOT Number", "Business Telephone No.", "State Incorporated",
        "Contact Telephone No.", "Duns & Bradstreet"
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name       = COLUMNS[cell.column - 1]
            cell.font      = Font(name="Calibri", color="000000", size=11)
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if col_name in CENTER_COLS else "left",
                vertical="center"
            )

    for ci, col_name in enumerate(COLUMNS, 1):
        letter  = get_column_letter(ci)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    try:
        wb.save(filepath)
        log(f"[✓] Saved: {filepath}")
    except PermissionError:
        ts        = int(time.time())
        base, ext = os.path.splitext(filepath)
        fallback  = f"{base}_{ts}{ext}"
        wb.save(fallback)
        log(f"[⚠] File locked — fallback: {fallback}")
        messagebox.showwarning("File Locked",
            f"The target file was open in Excel.\nFallback saved:\n{fallback}")


# ─── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("USDOT Intelligence Scraper — TTM Dispatch")
        self.geometry("860x720")
        self.resizable(True, True)
        self.configure(bg="#2b2b2b")
        self._build_ui()

    def _build_ui(self):
        pad = dict(padx=12, pady=6)

        tk.Label(self, text="⚡ USDOT Intelligence Scraper",
                 font=("Segoe UI", 16, "bold"), fg="#4fc3f7", bg="#2b2b2b"
                 ).pack(pady=(14, 2))
        tk.Label(self, text="TTM Dispatch Automation Suite",
                 font=("Segoe UI", 9), fg="#888888", bg="#2b2b2b"
                 ).pack(pady=(0, 10))

        cfg = tk.LabelFrame(self, text=" Output Configuration ",
                            font=("Segoe UI", 9, "bold"), fg="#aaaaaa",
                            bg="#2b2b2b", bd=1, relief="groove")
        cfg.pack(fill="x", **pad)

        row0 = tk.Frame(cfg, bg="#2b2b2b")
        row0.pack(fill="x", padx=8, pady=6)

        tk.Label(row0, text="Filename:", fg="#cccccc", bg="#2b2b2b",
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0,6))
        self.filename_var = tk.StringVar(value="DOT_Business_Data.xlsx")
        tk.Entry(row0, textvariable=self.filename_var, width=32,
                 bg="#3c3c3c", fg="#ffffff", insertbackground="white",
                 relief="flat", font=("Consolas", 9)
                 ).grid(row=0, column=1, sticky="w")

        tk.Label(row0, text="  Dir:", fg="#cccccc", bg="#2b2b2b",
                 font=("Segoe UI", 9)).grid(row=0, column=2, sticky="w", padx=(12,6))
        self.dir_var = tk.StringVar(value=os.path.expanduser("~\\Desktop"))
        tk.Entry(row0, textvariable=self.dir_var, width=26,
                 bg="#3c3c3c", fg="#ffffff", insertbackground="white",
                 relief="flat", font=("Consolas", 9)
                 ).grid(row=0, column=3, sticky="w")
        tk.Button(row0, text="Browse…", command=self._browse,
                  bg="#1F4E78", fg="white", relief="flat",
                  font=("Segoe UI", 8), cursor="hand2",
                  activebackground="#2d6fa3"
                  ).grid(row=0, column=4, padx=(6,0))

        self.headless_var = tk.BooleanVar(value=True)
        tk.Checkbutton(row0, text="Headless", variable=self.headless_var,
                       fg="#cccccc", bg="#2b2b2b", selectcolor="#3c3c3c",
                       activebackground="#2b2b2b", activeforeground="white",
                       font=("Segoe UI", 9)
                       ).grid(row=0, column=5, padx=(16,0))

        inp = tk.LabelFrame(self, text=" DOT Numbers (one per line) ",
                            font=("Segoe UI", 9, "bold"), fg="#aaaaaa",
                            bg="#2b2b2b", bd=1, relief="groove")
        inp.pack(fill="both", expand=False, **pad)

        self.dot_input = scrolledtext.ScrolledText(
            inp, height=8, bg="#1e1e1e", fg="#d4d4d4",
            insertbackground="white", font=("Consolas", 10),
            relief="flat", wrap="none"
        )
        self.dot_input.pack(fill="both", expand=True, padx=8, pady=6)
        self.dot_input.insert("1.0", "4581886\n")

        self.start_btn = tk.Button(
            self, text="⚡  Start Scraping Pipeline",
            command=self._launch,
            bg="#1F4E78", fg="white", activebackground="#2d6fa3",
            font=("Segoe UI", 11, "bold"), relief="flat",
            cursor="hand2", pady=8
        )
        self.start_btn.pack(fill="x", padx=12, pady=(4, 2))

        log_frame = tk.LabelFrame(self, text=" System Activity Log ",
                                  font=("Segoe UI", 9, "bold"), fg="#aaaaaa",
                                  bg="#2b2b2b", bd=1, relief="groove")
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_box = scrolledtext.ScrolledText(
            log_frame, bg="#1e1e1e", fg="#d4d4d4",
            insertbackground="white", font=("Consolas", 9),
            relief="flat", state="disabled"
        )
        self.log_box.pack(fill="both", expand=True, padx=8, pady=6)
        self.log_box.tag_config("ok",   foreground="#4ec9b0")
        self.log_box.tag_config("warn", foreground="#dcdcaa")
        self.log_box.tag_config("err",  foreground="#f48771")
        self.log_box.tag_config("info", foreground="#9cdcfe")

    def _browse(self):
        d = filedialog.askdirectory(initialdir=self.dir_var.get())
        if d:
            self.dir_var.set(d)

    def _log(self, msg):
        def _write():
            self.log_box.config(state="normal")
            ts   = datetime.now().strftime("%H:%M:%S")
            line = f"[{ts}] {msg}\n"
            tag  = "info"
            if any(x in msg for x in ["[✓]", "[✅]"]):  tag = "ok"
            elif any(x in msg for x in ["[⚠]", "[!]"]): tag = "warn"
            elif "[✗]" in msg:                           tag = "err"
            self.log_box.insert("end", line, tag)
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _write)

    def _launch(self):
        raw  = self.dot_input.get("1.0", "end").strip()
        dots = [d.strip() for d in raw.splitlines() if d.strip()]
        if not dots:
            messagebox.showwarning("No Input", "Enter at least one DOT number.")
            return

        fname = self.filename_var.get().strip()
        if not fname.endswith(".xlsx"):
            fname += ".xlsx"
        filepath = os.path.join(self.dir_var.get(), fname)

        self.start_btn.config(state="disabled", text="⏳  Running…")
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

        def worker():
            try:
                run_pipeline(dots, filepath, self.headless_var.get(), self._log)
            finally:
                self.after(0, lambda: self.start_btn.config(
                    state="normal", text="⚡  Start Scraping Pipeline"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    App().mainloop()
import streamlit as st
import time
import os
import re
from datetime import datetime
from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ────────────────────────────────────────────────────────────────
BASE_URL = "https://motus.dot.gov/customer/{}/account"
COLUMNS = [
    "DOT Number", "Legal Business Name", "Doing Business As (DBA)",
    "Principal Place of Business", "Mailing Address", "Business Telephone No.",
    "Duns & Bradstreet", "Form of Business", "State Incorporated",
    "Business Email", "Official Name", "Title",
    "Contact Telephone No.", "Contact Email"
]
HEADER_BG = "1F4E78"
HEADER_FG = "FFFFFF"
GRID_COLOR = "D9D9D9"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ─── Known field labels exactly as they appear on the page ───────────────────
BUSINESS_FIELDS = [
    ("legal_name", "Legal Business Name"),
    ("dba", "Doing Business As (DBA) Name"),
    ("principal", "Principal Place of Business"),
    ("mailing", "Mailing Address"),
    ("phone", "Business Telephone No."),
    ("duns", "Duns & Bradstreet"),
    ("form", "Form of Business"),
    ("state_inc", "State Incorporated"),
    ("email", "Business Email"),
]

OFFICIALS_HEADERS = ["Official Name", "Title", "Telephone No", "Email"]

# ─── Page text utilities ──────────────────────────────────────────────────────

def get_lines(page):
    """Return cleaned, non-empty lines from the full page innerText."""
    raw = page.evaluate("() => document.body.innerText")
    return [ln.strip() for ln in raw.split("\n") if ln.strip()]


def extract_business_fields(lines, log):
    """
    Walk lines sequentially. When we see a known label, the NEXT line
    that is NOT itself a known label is the value.
    """
    all_labels_lower = {lbl.lower() for _, lbl in BUSINESS_FIELDS}
    officials_lower = {h.lower() for h in OFFICIALS_HEADERS}
    skip_lower = all_labels_lower | officials_lower

    label_lookup = {lbl.lower(): key for key, lbl in BUSINESS_FIELDS}

    result = {}
    i = 0
    while i < len(lines):
        ll = lines[i].lower()
        if ll in label_lookup:
            key = label_lookup[ll]
            value = ""
            for j in range(i + 1, min(i + 4, len(lines))):
                candidate = lines[j]
                if candidate.lower() in skip_lower:
                    break
                value = candidate
                break
            result[key] = value
        i += 1

    log(f"[✦] Business fields: {result}")
    return result


def extract_officials(lines, log):
    """Find officials data blocks and read consecutive values."""
    hdr_lower = [h.lower() for h in OFFICIALS_HEADERS]
    n = len(hdr_lower)

    block_start = None
    for i in range(len(lines) - n + 1):
        window = [lines[i + k].lower() for k in range(n)]
        if window == hdr_lower:
            block_start = i
            break

    if block_start is None:
        for i, line in enumerate(lines):
            if line.lower() == hdr_lower[0]:
                block_start = i
                break

    if block_start is None:
        log("[Officials] ⚠ Header block not found in page text")
        return {"name": "N/A", "title": "N/A", "phone": "N/A", "email": "N/A"}

    val_start = block_start + n
    log(f"[Officials] Header block at line {block_start}, values at {val_start}")
    log(f"[Officials] Value lines: {lines[val_start:val_start+6]}")

    def safe_val(offset):
        idx = val_start + offset
        if idx < len(lines):
            v = lines[idx]
            if re.search(r"row(s)? (per page|selected)|^\d+[-–]\d+", v.lower()):
                return "N/A"
            return v if v else "N/A"
        return "N/A"

    result = {
        "name":  safe_val(0),
        "title": safe_val(1),
        "phone": safe_val(2),
        "email": safe_val(3),
    }
    log(f"[Officials] ✓ {result}")
    return result


def ensure_officials_open(page, log):
    """Click the Company Officials accordion only if it is currently closed."""
    try:
        panels = page.query_selector_all("mat-expansion-panel")
        for panel in panels:
            header = panel.query_selector("mat-expansion-panel-header")
            if not header:
                continue
            if "official" not in header.inner_text().lower():
                continue
            cls  = header.get_attribute("class") or ""
            aria = header.get_attribute("aria-expanded") or ""
            if "mat-expanded" in cls or aria == "true":
                log("[✓] Officials panel already open")
            else:
                log("[↕] Opening officials panel")
                header.click()
                time.sleep(1.5)
            return
        header = page.query_selector("mat-expansion-panel-header")
        if header:
            cls  = header.get_attribute("class") or ""
            aria = header.get_attribute("aria-expanded") or ""
            if "mat-expanded" not in cls and aria != "true":
                log("[↕] Opening first expansion panel")
                header.click()
                time.sleep(1.5)
    except Exception as e:
        log(f"[!] Panel check error: {e}")


# ─── Main scrape ──────────────────────────────────────────────────────────────

def scrape_dot(page, dot, log):
    url = BASE_URL.format(dot)
    log(f"[→] {url}")
    page.goto(url, wait_until="networkidle", timeout=60000)
    time.sleep(2.5)

    ensure_officials_open(page, log)
    time.sleep(1)

    lines = get_lines(page)
    log(f"[✦] Total lines captured: {len(lines)}")

    biz      = extract_business_fields(lines, log)
    officials = extract_officials(lines, log)

    def g(key):
        v = biz.get(key, "")
        return v if v else "N/A"

    return {
        "DOT Number":                  dot,
        "Legal Business Name":         g("legal_name"),
        "Doing Business As (DBA)":     g("dba"),
        "Principal Place of Business": g("principal"),
        "Mailing Address":             g("mailing"),
        "Business Telephone No.":      g("phone"),
        "Duns & Bradstreet":           g("duns"),
        "Form of Business":            g("form"),
        "State Incorporated":          g("state_inc"),
        "Business Email":              g("email"),
        "Official Name":               officials["name"],
        "Title":                       officials["title"],
        "Contact Telephone No.":       officials["phone"],
        "Contact Email":               officials["email"],
    }


# ─── Pipeline ─────────────────────────────────────────────────────────────────

def run_pipeline(dot_numbers, filepath, log):
    records = []
    log(f"[⚡] Pipeline start — {len(dot_numbers)} DOT number(s)")

    with sync_playwright() as pw:
        # Optimized with standard flags to run inside Linux cloud instances smoothly
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]
        )
        ctx     = browser.new_context(user_agent=USER_AGENT)
        page    = ctx.new_page()

        for i, dot in enumerate(dot_numbers, 1):
            dot = dot.strip()
            if not dot:
                continue
            log(f"\n── [{i}/{len(dot_numbers)}] DOT {dot} ──")
            try:
                records.append(scrape_dot(page, dot, log))
                log(f"[✓] DOT {dot} complete")
            except Exception as e:
                log(f"[✗] DOT {dot} error: {e}")
                empty = {col: "N/A" for col in COLUMNS}
                empty["DOT Number"] = dot
                records.append(empty)

        browser.close()

    if records:
        write_excel(records, filepath, log)
        log(f"\n[✅] Done — {len(records)} record(s) exported.")
        return True
    else:
        log("[⚠] No records to export.")
        return False


# ─── Excel writer ─────────────────────────────────────────────────────────────

def write_excel(records, filepath, log):
    df = pd.DataFrame(records, columns=COLUMNS)
    df.to_excel(filepath, index=False, sheet_name="DOT Intelligence")

    wb = load_workbook(filepath)
    ws = wb.active

    thin   = Side(style="thin", color=GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
    hdr_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)

    for ci in range(1, len(COLUMNS) + 1):
        cell           = ws.cell(row=1, column=ci)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    CENTER_COLS = {
        "DOT Number", "Business Telephone No.", "State Incorporated",
        "Contact Telephone No.", "Duns & Bradstreet"
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name       = COLUMNS[cell.column - 1]
            cell.font      = Font(name="Calibri", color="000000", size=11)
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if col_name in CENTER_COLS else "left",
                vertical="center"
            )

    for ci, col_name in enumerate(COLUMNS, 1):
        letter  = get_column_letter(ci)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(filepath)
    log(f"[✓] Formatted output generated.")


# ─── Streamlit Web UI Layout ──────────────────────────────────────────────────

st.set_page_config(
    page_title="TTM Engine | USDOT Scraper",
    page_icon="🚛",
    layout="wide"
)

# High-Contrast Professional Custom Styling Injector
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;700&display=swap');
    .main { background-color: #0f111a; color: #ffffff; font-family: 'Poppins', sans-serif; }
    h1, h2, h3 { color: #ff3333 !important; font-family: 'Poppins', sans-serif; font-weight: 700; }
    .stButton>button {
        background-color: #1F4E78 !important; color: white !important;
        font-weight: bold; border-radius: 4px; border: none; width: 100%; height: 45px; transition: 0.3s;
    }
    .stButton>button:hover { background-color: #2d6fa3 !important; }
    div[data-testid="stFrameMutedContainer"] { background-color: #1a1c24 !important; border: 1px solid #333; }
    iframe { background-color: #1e1e1e !important; }
    </style>
""", unsafe_allow_html=True)

st.title("⚡ USDOT Intelligence Scraper")
st.caption("TTM Dispatch Automation Suite — Cloud Terminal Instance")
st.markdown("---")

# UI Configuration Split Setup
col_left, col_right = st.columns([2, 1])

with col_right:
    st.markdown("### ⚙️ Output Configuration")
    filename_input = st.text_input("Output Filename:", value="DOT_Business_Data.xlsx")
    if not filename_input.endswith(".xlsx"):
        filename_input += ".xlsx"
    
    st.info("💡 **Cloud Environment Mode:** Files are written securely inside the server session, compile formatting rules dynamically, and return directly to your browser download directory instantly.")

with col_left:
    st.markdown("### 📋 Target Inputs")
    raw_dot_input = st.text_area("DOT Numbers (One entry per line text stream):", value="4581886", height=180)

st.markdown("---")

# Streamlit-compatible Logging Architecture Block
st.markdown("### 📜 System Activity Log")
log_container = st.empty()

# Persistent tracking initialization for updates
if "log_history" not in st.session_state:
    st.session_state.log_history = []

def web_log_writer(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    formatted_line = f"[{ts}] {msg}"
    st.session_state.log_history.append(formatted_line)
    # Output directly inside scrolling window
    log_container.code("\n".join(st.session_state.log_history), language="text")

# Render historical records upon layout state shifts
if st.session_state.log_history:
    log_container.code("\n".join(st.session_state.log_history), language="text")

# Trigger Action Routine Block
if st.button("⚡   Start Scraping Pipeline"):
    dots_list = [d.strip() for d in raw_dot_input.splitlines() if d.strip()]
    
    if not dots_list:
        st.error("Missing Parameter: Please supply one or more DOT sequence items before initiating run pipelines.")
    else:
        st.session_state.log_history = [] # Reset old logs on new run
        temp_filepath = "cloud_output.xlsx"
        
        # Execute underlying synchronous target sequence steps
        success = run_pipeline(dots_list, temp_filepath, web_log_writer)
        
        if success and os.path.exists(temp_filepath):
            with open(temp_filepath, "rb") as f:
                excel_bytes = f.read()
            
            st.markdown("---")
            st.success("🎉 Data capture sequence finalized cleanly!")
            st.download_button(
                label="📥 DOWNLOAD FORMATTED DOT INTELLIGENCE SHEET",
                data=excel_bytes,
                file_name=filename_input,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Clean server memory temp files
            try:
                os.remove(temp_filepath)
            except:
                pass
